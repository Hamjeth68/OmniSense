
from typing import Dict, List, Optional, Any, Tuple
import logging
import asyncio
import re
import os
import tempfile
from pathlib import Path
import json

# Import HuggingFace components
import torch
from transformers import (
    AutoTokenizer, AutoModelForSeq2SeqLM,
    AutoModelForSequenceClassification,
    AutoModelForQuestionAnswering,
    pipeline,
    BertTokenizer, BertModel
)

# Import document processing libraries
import PyPDF2
import docx
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl

# Import NLP libraries
import spacy
from sentence_transformers import SentenceTransformer
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from nltk.chunk import ne_chunk
from nltk.tag import pos_tag

# Download required NLTK data
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt')

try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords')

try:
    nltk.data.find('taggers/averaged_perceptron_tagger')
except LookupError:
    nltk.download('averaged_perceptron_tagger')

try:
    nltk.data.find('chunkers/maxent_ne_chunker')
except LookupError:
    nltk.download('maxent_ne_chunker')

try:
    nltk.data.find('corpora/words')
except LookupError:
    nltk.download('words')

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TextProcessor:
    """
    Processes text content for summarization, Q&A, topic extraction,
    sentiment analysis, and other NLP tasks.
    """
    
    def __init__(self):
        """Initialize the text processor with required models"""
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        
        # Model names
        self.summarization_model_name = "facebook/bart-large-cnn"
        self.qa_model_name = "deepset/roberta-base-squad2"
        self.sentiment_model_name = "cardiffnlp/twitter-roberta-base-sentiment-latest"
        self.ner_model_name = "dbmdz/bert-large-cased-finetuned-conll03-english"
        self.embedding_model_name = "sentence-transformers/all-MiniLM-L6-v2"
        
        # Initialize models (lazy loading)
        self.summarizer = None
        self.qa_pipeline = None
        self.sentiment_pipeline = None
        self.ner_pipeline = None
        self.embedding_model = None
        
        # Load spaCy model
        try:
            self.nlp = spacy.load("en_core_web_sm")
        except OSError:
            logger.warning("spaCy model not found. Run: python -m spacy download en_core_web_sm")
            self.nlp = None
        
        logger.info(f"Text processor initialized (device: {self.device})")
    
    async def _ensure_models_loaded(self):
        """Ensure all required models are loaded"""
        if self.summarizer is None:
            logger.info("Loading summarization model...")
            self.summarizer = pipeline(
                "summarization",
                model=self.summarization_model_name,
                device=0 if self.device == "cuda" else -1
            )
        
        if self.qa_pipeline is None:
            logger.info("Loading Q&A model...")
            self.qa_pipeline = pipeline(
                "question-answering",
                model=self.qa_model_name,
                device=0 if self.device == "cuda" else -1
            )
        
        if self.sentiment_pipeline is None:
            logger.info("Loading sentiment analysis model...")
            self.sentiment_pipeline = pipeline(
                "sentiment-analysis",
                model=self.sentiment_model_name,
                device=0 if self.device == "cuda" else -1
            )
        
        if self.ner_pipeline is None:
            logger.info("Loading NER model...")
            self.ner_pipeline = pipeline(
                "ner",
                model=self.ner_model_name,
                aggregation_strategy="simple",
                device=0 if self.device == "cuda" else -1
            )
        
        if self.embedding_model is None:
            logger.info("Loading embedding model...")
            self.embedding_model = SentenceTransformer(self.embedding_model_name)
    
    async def extract_text_from_document(self, file_path: str) -> str:
        """
        Extract text content from various document formats
        
        Args:
            file_path: Path to document file
            
        Returns:
            Extracted text content
        """
        file_extension = Path(file_path).suffix.lower()
        
        try:
            if file_extension == ".pdf":
                return await self._extract_from_pdf(file_path)
            elif file_extension in [".doc", ".docx"]:
                return await self._extract_from_docx(file_path)
            elif file_extension in [".xls", ".xlsx"]:
                return await self._extract_from_excel(file_path)
            elif file_extension == ".txt":
                return await self._extract_from_txt(file_path)
            elif file_extension in [".html", ".htm"]:
                return await self._extract_from_html(file_path)
            else:
                logger.warning(f"Unsupported file format: {file_extension}")
                return ""
        except Exception as e:
            logger.error(f"Text extraction error: {str(e)}")
            return ""
    
    async def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text = ""
        with open(file_path, "rb") as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text.strip()
    
    async def _extract_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        doc = docx.Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    
    async def _extract_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        try:
            # Try with pandas first
            df = pd.read_excel(file_path, sheet_name=None)
            text = ""
            for sheet_name, sheet_df in df.items():
                text += f"Sheet: {sheet_name}\n"
                text += sheet_df.to_string() + "\n\n"
            return text.strip()
        except Exception:
            # Fallback to openpyxl
            wb = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell else "" for cell in row])
                    text += row_text + "\n"
                text += "\n"
            return text.strip()
    
    async def _extract_from_txt(self, file_path: str) -> str:
        """Extract text from TXT file"""
        with open(file_path, "r", encoding="utf-8") as file:
            return file.read()
    
    async def _extract_from_html(self, file_path: str) -> str:
        """Extract text from HTML file"""
        with open(file_path, "r", encoding="utf-8") as file:
            soup = BeautifulSoup(file.read(), "html.parser")
            return soup.get_text()
    
    async def summarize(self, text: str, max_length: int = 150, min_length: int = 50) -> str:
        """
        Generate a summary of the text
        
        Args:
            text: Input text to summarize
            max_length: Maximum length of summary
            min_length: Minimum length of summary
            
        Returns:
            Generated summary
        """
        await self._ensure_models_loaded()
        
        try:
            # Split text into chunks if too long
            max_chunk_length = 1024  # BART's max input length
            
            if len(text.split()) <= max_chunk_length:
                # Single chunk
                summary = self.summarizer(
                    text,
                    max_length=max_length,
                    min_length=min_length,
                    do_sample=False
                )[0]["summary_text"]
            else:
                # Multiple chunks
                sentences = sent_tokenize(text)
                chunks = []
                current_chunk = ""
                
                for sentence in sentences:
                    if len((current_chunk + " " + sentence).split()) <= max_chunk_length:
                        current_chunk += " " + sentence
                    else:
                        if current_chunk:
                            chunks.append(current_chunk.strip())
                        current_chunk = sentence
                
                if current_chunk:
                    chunks.append(current_chunk.strip())
                
                # Summarize each chunk
                chunk_summaries = []
                for chunk in chunks:
                    try:
                        chunk_summary = self.summarizer(
                            chunk,
                            max_length=max_length // len(chunks) + 20,
                            min_length=min_length // len(chunks),
                            do_sample=False
                        )[0]["summary_text"]
                        chunk_summaries.append(chunk_summary)
                    except Exception as e:
                        logger.warning(f"Failed to summarize chunk: {str(e)}")
                        continue
                
                # Combine chunk summaries
                combined_summary = " ".join(chunk_summaries)
                
                # Final summarization if needed
                if len(combined_summary.split()) > max_length:
                    summary = self.summarizer(
                        combined_summary,
                        max_length=max_length,
                        min_length=min_length,
                        do_sample=False
                    )[0]["summary_text"]
                else:
                    summary = combined_summary
            
            logger.info(f"Summary generated: {len(summary)} characters")
            return summary
            
        except Exception as e:
            logger.error(f"Summarization error: {str(e)}")
            # Fallback to extractive summarization
            return await self._extractive_summarization(text, max_length)
    
    async def _extractive_summarization(self, text: str, max_length: int) -> str:
        """Fallback extractive summarization"""
        sentences = sent_tokenize(text)
        if len(sentences) <= 3:
            return text
        
        # Simple extractive approach: take first, middle, and important sentences
        important_sentences = sentences[:2] + sentences[len(sentences)//2:len(sentences)//2+1] + sentences[-1:]
        summary = " ".join(important_sentences[:max_length//20])  # Rough word count estimate
        return summary
    
    async def extract_action_items(self, text: str) -> List[Dict[str, Any]]:
        """
        Extract action items from text
        
        Args:
            text: Input text
            
        Returns:
            List of action items with details
        """
        await self._ensure_models_loaded()
        
        try:
            # Patterns for action items
            action_patterns = [
                r"(?:TODO|To do|Action item|Task|Follow up|Next step)[:.]?\s*(.+?)(?:\n|$)",
                r"(?:^|\n)\s*[-*•]\s*(.+?(?:will|should|must|need to|has to).+?)(?:\n|$)",
                r"(?:I|We|You|They|[A-Z][a-z]+)\s+(?:will|should|must|need to|has to)\s+(.+?)(?:\.|!|\?|$)",
                r"@(\w+)\s+(.+?)(?:\n|$)",  # Mentions with tasks
            ]
            
            action_items = []
            
            for pattern in action_patterns:
                matches = re.finditer(pattern, text, re.MULTILINE | re.IGNORECASE)
                for match in matches:
                    action_text = match.group(1).strip()
                    if len(action_text) > 10:  # Filter out very short matches
                        
                        # Extract assignee if mentioned
                        assignee = None
                        assignee_patterns = [
                            r"@(\w+)",
                            r"(?:assign(?:ed)? to|for) (\w+)",
                            r"(\w+) (?:will|should|must)"
                        ]
                        
                        for ap in assignee_patterns:
                            assignee_match = re.search(ap, action_text, re.IGNORECASE)
                            if assignee_match:
                                assignee = assignee_match.group(1)
                                break
                        
                        # Extract due date if mentioned
                        due_date = None
                        date_patterns = [
                            r"by (\w+ \d+)",
                            r"due (\w+ \d+)",
                            r"before (\w+ \d+)",
                            r"(\d{1,2}/\d{1,2})",
                            r"(next week|this week|tomorrow|today)"
                        ]
                        
                        for dp in date_patterns:
                            date_match = re.search(dp, action_text, re.IGNORECASE)
                            if date_match:
                                due_date = date_match.group(1)
                                break
                        
                        # Extract priority
                        priority = "medium"  # default
                        if re.search(r"urgent|asap|immediately|high priority", action_text, re.IGNORECASE):
                            priority = "high"
                        elif re.search(r"low priority|when possible|eventually", action_text, re.IGNORECASE):
                            priority = "low"
                        
                        action_items.append({
                            "text": action_text,
                            "assignee": assignee,
                            "due_date": due_date,
                            "priority": priority,
                            "status": "pending"
                        })
            
            # Remove duplicates
            unique_actions = []
            seen_texts = set()
            for action in action_items:
                if action["text"] not in seen_texts:
                    unique_actions.append(action)
                    seen_texts.add(action["text"])
            
            logger.info(f"Extracted {len(unique_actions)} action items")
            return unique_actions
            
        except Exception as e:
            logger.error(f"Action item extraction error: {str(e)}")
            return []
    
    async def extract_key_points(self, text: str) -> List[str]:
        """
        Extract key points from text
        
        Args:
            text: Input text
            
        Returns:
            List of key points
        """
        try:
            sentences = sent_tokenize(text)
            
            # Score sentences based on various factors
            sentence_scores = {}
            
            # Get word frequencies
            words = word_tokenize(text.lower())
            stop_words = set(stopwords.words('english'))
            words = [word for word in words if word.isalnum() and word not in stop_words]
            
            word_freq = {}
            for word in words:
                word_freq[word] = word_freq.get(word, 0) + 1
            
            # Score sentences
            for sentence in sentences:
                sentence_words = word_tokenize(sentence.lower())
                sentence_words = [word for word in sentence_words if word.isalnum() and word not in stop_words]
                
                score = 0
                for word in sentence_words:
                    score += word_freq.get(word, 0)
                
                # Bonus for sentences with numbers, dates, names
                if re.search(r'\d+', sentence):
                    score += 2
                if re.search(r'[A-Z][a-z]+ [A-Z][a-z]+', sentence):  # Names
                    score += 1
                
                # Penalty for very short or very long sentences
                if len(sentence.split()) < 5:
                    score *= 0.5
                elif len(sentence.split()) > 30:
                    score *= 0.7
                
                sentence_scores[sentence] = score
            
            # Get top sentences
            sorted_sentences = sorted(sentence_scores.items(), key=lambda x: x[1], reverse=True)
            
            # Take top 30% of sentences, but at least 3 and at most 10
            num_key_points = max(3, min(10, len(sentences) // 3))
            key_points = [sentence for sentence, score in sorted_sentences[:num_key_points]]
            
            logger.info(f"Extracted {len(key_points)} key points")
            return key_points
            
        except Exception as e:
            logger.error(f"Key point extraction error: {str(e)}")
            return sentences[:5] if sentences else []
    
    async def extract_structured_data(self, file_path: str) -> Dict[str, Any]:
        """
        Extract structured data (tables, charts) from documents
        
        Args:
            file_path: Path to document file
            
        Returns:
            Dict containing structured data
        """
        file_extension = Path(file_path).suffix.lower()
        structured_data = {
            "tables": [],
            "charts": [],
            "lists": [],
            "metadata": {}
        }
        
        try:
            if file_extension in [".xls", ".xlsx"]:
                # Extract Excel data
                wb = openpyxl.load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    # Extract table data
                    table_data = []
                    for row in sheet.iter_rows(values_only=True):
                        row_data = [str(cell) if cell else "" for cell in row]
                        if any(row_data):  # Skip empty rows
                            table_data.append(row_data)
                    
                    if table_data:
                        structured_data["tables"].append({
                            "sheet_name": sheet_name,
                            "data": table_data,
                            "rows": len(table_data),
                            "columns": len(table_data[0]) if table_data else 0
                        })
            
            elif file_extension == ".pdf":
                # For PDF, we'd need more sophisticated table extraction
                # This is a simplified version
                text = await self._extract_from_pdf(file_path)
                
                # Extract lists
                list_items = re.findall(r'^\s*[-•*]\s*(.+)$', text, re.MULTILINE)
                if list_items:
                    structured_data["lists"].append({
                        "type": "bullet_list",
                        "items": list_items
                    })
                
                # Extract numbered lists
                numbered_items = re.findall(r'^\s*\d+[.)]\s*(.+)$', text, re.MULTILINE)
                if numbered_items:
                    structured_data["lists"].append({
                        "type": "numbered_list",
                        "items": numbered_items
                    })
            
            logger.info(f"Extracted structured data: {len(structured_data['tables'])} tables, {len(structured_data['lists'])} lists")
            return structured_data
            
        except Exception as e:
            logger.error(f"Structured data extraction error: {str(e)}")
            return structured_data
    
    async def create_qa_model(self, text: str, file_path: str) -> Dict[str, Any]:
        """
        Create a Q&A model for the document